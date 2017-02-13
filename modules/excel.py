import openpyxl

KINTERA_FIRST_NAME_COL = 6
KINTERA_LAST_NAME_COL = 5
KINTERA_GIFT_CODE_COL = 7
KINTERA_SCG_GIFT_CODE = 'SCGM |'

GENERAL_NAME_COL = 1
GENERAL_EMAIL_COL = 3
GENERAL_GIFT_CODE_COL = 9
GENERAL_SCG_GIFT_CODES = ["Odyssey Unclassified TBD", "College Fund", "Dean's Fund for Student Life", "Jeff Metcalf Internships"]

class giving_document(object):

    def __init__(self, filename):
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb.worksheets[0]
        self.type = self.check_sheet_type()

    def check_sheet_type(self):
        """Returns the type of giving document
        Returns:
            (str): 'KINTERA', 'GENERAL, 'OTHER'
        """
        if self.sheet['A1'].value == 'Date Entered':
            return 'KINTERA'
        if self.sheet['A1'].value == 'Entity ID':
            return 'GENERAL'
        else:
            return 'OTHER'

    def get_donors(self):
        """Gets the applicable SCG donors from the doc"""

        donors = []
        if self.type == 'KINTERA':
            for row in self.sheet.iter_rows(min_row=2):
                if row[KINTERA_GIFT_CODE_COL].value == KINTERA_SCG_GIFT_CODE:
                    fname = row[KINTERA_FIRST_NAME_COL].value
                    lname = row[KINTERA_LAST_NAME_COL].value
                    donor = {'fname': fname, 'lname': lname, 'email': ''}
                    donors.append(donor)
        elif self.type == 'GENERAL':
            for row in self.sheet.iter_rows(min_row=2):
                if row[GENERAL_GIFT_CODE_COL].value in GENERAL_SCG_GIFT_CODES:
                    name = row[GENERAL_NAME_COL].value
                    lname = name.split(',')[0]
                    fname = name.split(' ')[1]
                    email = row[GENERAL_EMAIL_COL].value
                    donor = {'fname': fname, 'lname': lname, 'email': email}
                    donors.append(donor)
        else:
            pass
        return donors



if __name__ == "__main__":
    doc = giving_document('sample_general.xlsx')
    for p in doc.get_donors():
        print "%s %s %s" % (p['fname'], p['lname'], p['email'])
